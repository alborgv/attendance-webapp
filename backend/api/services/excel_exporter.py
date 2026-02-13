import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


HEADER_FILL = PatternFill(
    start_color="1F4ED8",
    end_color="1F4ED8",
    fill_type="solid"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

HEADER_ALIGNMENT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=False
)

TITLE_FONT = Font(
    bold=True,
    size=16
)

TITLE_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center"
)


def export_to_excel(
    *,
    filename: str,
    title: str | None = None,
    columns: list[dict],
    data: list[dict],
):
    """
    columns: [
        {"key": "nombre_completo", "label": "Nombre completo"},
        ...
    ]
    data: list[dict]
    """

    df = pd.DataFrame(data)
    ordered_columns = [c["key"] for c in columns]
    df = df[ordered_columns]
    df.columns = [c["label"].upper() for c in columns]

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'

    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        start_row = 1 if title else 0

        df.to_excel(
            writer,
            index=False,
            sheet_name="Datos",
            startrow=start_row
        )

        sheet = writer.book.active

        total_columns = len(df.columns)

        # TITULO
        if title:
            sheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=total_columns
            )
            title_cell = sheet.cell(row=1, column=1)
            title_cell.value = title.upper()
            title_cell.font = TITLE_FONT
            title_cell.alignment = TITLE_ALIGNMENT

        header_row = 2 if title else 1

        # ENCABEZADO
        for col_idx, cell in enumerate(sheet[header_row], start=1):
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

        # AJUSTE DE ANCHO (header + data)
        for col_idx, column_name in enumerate(df.columns, start=1):
            column_letter = get_column_letter(col_idx)

            max_length = len(str(column_name))
            for row in range(header_row + 1, sheet.max_row + 1):
                cell_value = sheet[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            sheet.column_dimensions[column_letter].width = min(max_length + 6, 45)

        # WRAP TEXT PARA DATOS
        for row in sheet.iter_rows(
            min_row=header_row + 1,
            max_row=sheet.max_row
        ):
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

    return response
